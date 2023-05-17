# Import necessary libraries
from xml.dom.minidom import parse
import xml.dom.minidom
import openpyxl  # Write Excel

# Creat an Excel
workBook = openpyxl.Workbook()
sheet = workBook.active
sheet.title = "autophagosome"

title = ["id","name","definition","childnodes"]
for q in range(len(title)):
  sheet.cell(1,q+1,title[q])

# Read the data
r = 2
tree = xml.dom.minidom.parse("go_obo.xml")
root = tree.documentElement
terms = root.getElementsByTagName("term")
for term in terms:
  count = 0
  idList = []
  defstr = term.getElementsByTagName("defstr")[0].childNodes[0].data
  if "autophagosome" in defstr:
    id = term.getElementsByTagName("id")[0].childNodes[0].data
    name = term.getElementsByTagName("name")[0].childNodes[0].data
    for new_term in terms:
      new_ids = new_term.getElementsByTagName("id")[0].childNodes[0].data
      is_a = new_term.getElementsByTagName("is_a")
      for i in range(len(is_a)):
        if id in is_a[i].childNodes[0].data:
          count += 1
          idList.append(new_ids)
          for new_id in idList: 
            for newer_term in terms: 
              newer_id = newer_term.getElementsByTagName("id")[0].childNodes[0].data
              new_is_a = newer_term.getElementsByTagName("is_a")
              for t in range(len(new_is_a)):
                if new_id in new_is_a[t].childNodes[0].data:
                  count += 1
                  idList.append(newer_id)

    # Import date
    data = [id,name,defstr,count]
    for j in range(len(data)):
      sheet.cell(r,j+1,data[j])
    r += 1
# Save workbook
workBook.save(filename="autophagosome.xlsx")
